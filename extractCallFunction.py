import os
import re
import datetime
import csv
import shutil
from concurrent.futures import ProcessPoolExecutor
import concurrent
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from collections import OrderedDict

#current datetime
dt_now = datetime.datetime.now()
crrDir = os.path.dirname(__file__)

importList = []
importList_header = []
importList_detail = []

SQLID_COL = 17
SQL_COL = 18

# キーワードリストを直接指定
SQL_TARGETS_LIST = [
    "(+)",
    "CONCAT",
    "CURRENT DATE",
    "CURRENT TIME",
    "CURRENT TIMESTAMP",
    "DATE(",
    "DECODE",
    "DECODE(",
    "FETCH FIRST ",
    "FROM (",
    "HOUR(",
    "INT(",
    "INTEGER",
    "INTEGER(",
    "MINUTE",
    "MINUTE(",
    "NVL",
    "OPTIMIZE FOR",
    "REPLACE",
    "ROUND",
    "ROWNUM",
    "SUBSTR",
    "TO_CHAR",
    "TO_DATE",
    "TRIM",
    "TRUNC",
    "VALUE(",
    "||"
]

def getTimeString():
    """get Month
    Args:
    """        
    return dt_now.strftime('%Y%m%d%H%M%S')

def load_csv_to_objects(file_path):
    objects = []
    
    # CSVファイルを一括ロード
    with open(file_path, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)  # ヘッダーをキーとした辞書形式で各行を取得
        
        # 各行を辞書形式でリストに格納
        for row in reader:
            obj = {}
            for header, value in row.items():
                keys = header.split('.')  # 'object.item1' のようなドット区切りのキーを分割
                current_obj = obj
                for key in keys[:-1]:  # ネストされた辞書を作成
                    if key not in current_obj:
                        current_obj[key] = {}
                    current_obj = current_obj[key]
                current_obj[keys[-1]] = value  # 最後の要素に値を割り当て
            objects.append(obj)
    
    return (file_path,objects)


def extract_self_functions(line,function_pattern):
    """
    階層的な関数呼び出しから、クラス名と関数名をすべて抽出する。

    Args:
        line (str): 関数呼び出しの文字列。

    Returns:
        list: 抽出された (クラス名, 関数名) のタプルのリスト。
    """
    # 正規表現: クラス名.関数名(任意の引数)
    regex = re.compile(rf"({function_pattern})\s*\(")
    matches = regex.findall(line)

    return matches

def extract_functions(line,function_pattern):
    """
    階層的な関数呼び出しから、クラス名と関数名をすべて抽出する。

    Args:
        line (str): 関数呼び出しの文字列。

    Returns:
        list: 抽出された (クラス名, 関数名) のタプルのリスト。
    """
    # 正規表現: クラス名.関数名(任意の引数)
    regex = re.compile(rf"(\w+)\.({function_pattern})\(")
    matches = regex.findall(line)

    return matches

def extract_nested_functions(line,function_pattern):
    """
    階層的な関数呼び出しからすべてのクラス名と関数名を再帰的に抽出する。

    Args:
        line (str): 関数呼び出しの文字列。
        function_pattern(str):

    Returns:
        list: 抽出された (クラス名, 関数名) のタプルのリスト。
    """
    # 抽出結果
    results = []
    
    # 最外層の関数呼び出しを抽出
    matches = extract_functions(line,function_pattern)
    results.extend(matches)
    
    # 各マッチについて、引数部分を再解析
    for match in matches:
        # マッチ箇所を置き換えて引数部分を抽出
        function_call = f"{match[0]}.{match[1]}("
        start_index = line.find(function_call) + len(function_call)
        end_index = line.rfind(")")
        if start_index < end_index:
            nested_content = line[start_index:end_index]
            # 再帰的にネストされた部分を解析
            results.extend(extract_nested_functions(nested_content,function_pattern))
    
    return results

def parentWk(pair,processdict,parentPairs,line,matches,tmpFunctionList,line_number,callFunc,retDist):


        tmpChildFolder = pair["child_folder"]
        tmpChildPath = pair["child_path"]

        tmpParentFunctionList=[]                
        tmpParentFunctionList = [item for item in processdict[tmpChildFolder] if item["fileNameFull"] == tmpChildPath]
        tmpParentfunction_pattern = "|".join(map(re.escape, [item['function'] for item in tmpParentFunctionList]))

        matches = extract_self_functions(line,tmpParentfunction_pattern)
        if len(matches) > 0:

            for match in matches:
                caller_function_name = ""
                callee_function_name = ""        
                callee_function_name = (match)

                #Check Caller Func
                for callerFunction in tmpFunctionList:                   
                    if int(callerFunction["startNum"]) <= line_number <= int(callerFunction["endNum"]):
                        if callerFunction['function'] == callee_function_name:
                            continue
                        caller_function_name = callerFunction['function']
                        break

                if match == "" or caller_function_name == "":
                    continue

                parentKey = str.upper(callFunc["fileNameFull"] + "_" + caller_function_name)
                childKey = str.upper(callFunc["fileNameFull"] + "_" + callee_function_name)
                
                if (parentKey,childKey) in retDist :
                    continue
                
                retDist[(parentKey,childKey)] = [caller_function_name+"_"+ callFunc['fileName'], callee_function_name+"_"+ callFunc['fileName'],False] #0:Function / 1:SQL
                print(parentKey+","+childKey+","+caller_function_name+"_"+ callFunc['fileName']+","+callee_function_name+"_"+ callFunc['fileName'])

        else:
            pairList = [item for item in parentPairs if item["parent_path"] == pair["child_path"]]
            if len(pairList) > 0:
                for pair in pairList:
                    parentWk(pair,processdict,parentPairs,line,matches,tmpFunctionList,line_number,callFunc,retDist)          
            else:
                return
    

def call(file_path,target,processdict,importList_header,importList_detail,importList_SQL,parent_path):

    print(f"Start call : {file_path}") 
    #callFunctions = load_csv_to_objects(file_path)

    retDist = {}
    tmpFileName = ""
    targetProcess = processdict[target]
    for callFunc in targetProcess:
              
        if tmpFileName == callFunc['fileName']:
            continue
        else:
            tmpFileName = callFunc['fileName'] 

        # if tmpFileName != "EDWCustomerInput":
        #    continue
        
        #Function-SQL
        filtered_SQL = [item for item in importList_SQL if str.upper(callFunc["fileNameFull"]) == str.upper(item["ParentPath"]+"\\"+item['fileName'])]

        for data_SQL in filtered_SQL:

            caller_function_name = ""
            #Check Caller Func
            for callerFunction in  [item for item in targetProcess if str.upper(item['fileName']) == str.upper(callFunc['fileName']) and str.upper(item["fileNameFull"]) == str.upper(callFunc["fileNameFull"])] :                   
                if int(callerFunction["startNum"]) <= int(data_SQL["colNum"]) <= int(callerFunction["endNum"]):
                    caller_function_name = callerFunction['function']
                    break

            if caller_function_name == "":
                continue
            
            parentKey = str.upper(callFunc["fileNameFull"] + "_" + caller_function_name)
            childKey = data_SQL['funcition'] + "_" + "SQL"
            
            if (parentKey,childKey) in retDist :
                continue
            
            retDist[(parentKey,childKey)] = [caller_function_name+"_"+ callFunc['fileName'], data_SQL['funcition'],True]             

        filtered_SQL.clear()

        #Function-Function(self)
        tmpFunctionList=[]                
        tmpFunctionList = [item for item in targetProcess if (str.upper(item["fileNameFull"]) == str.upper(callFunc["fileNameFull"]) )]
        function_pattern = "|".join(map(re.escape, [item['function'] for item in tmpFunctionList]))

        in_comment_scope = False
        parentPairs = load_csv_to_objects(parent_path)[1]

        with open(callFunc["fileNameFull"], 'r', encoding='utf-8') as file:
            for line_number, line in enumerate(file, 1):

                matches = extract_self_functions(line,function_pattern)            

                if not(in_comment_scope) and line.strip().startswith("/*") and not(line.strip().endswith("*/")):
                    in_comment_scope = True

                if in_comment_scope and line.strip().endswith("*/"):
                    in_comment_scope = False
                    continue

                if not(in_comment_scope) :

                    # if tmpFileName != "EDWCustomerInput":
                    #     continue
                    # else:
                    #     if line_number > 125:
                    #         print()

                    if len(matches) > 0:                        

                        for match in matches:
                            caller_function_name = ""
                            callee_function_name = ""        
                            callee_function_name = (match)
            
                            #Check Caller Func
                            for callerFunction in tmpFunctionList:                   
                                if int(callerFunction["startNum"]) <= line_number <= int(callerFunction["endNum"]):
                                    if callerFunction['function'] == callee_function_name:
                                        continue
                                    caller_function_name = callerFunction['function']
                                    break

                            if caller_function_name == "":
                                continue
                            
                            parentKey = str.upper(callFunc["fileNameFull"] + "_" + caller_function_name)
                            childKey = str.upper(callFunc["fileNameFull"] + "_" + callee_function_name)
                            
                            if (parentKey,childKey) in retDist :
                                continue
                            
                            retDist[(parentKey,childKey)] = [caller_function_name+"_"+ callFunc['fileName'], callee_function_name+"_"+ callFunc['fileName'],False] #0:Function / 1:SQL
                            print(parentKey+","+childKey+","+caller_function_name+"_"+ callFunc['fileName']+","+callee_function_name+"_"+ callFunc['fileName'])

                    else:

                        pairList = [item for item in parentPairs if callFunc["fileNameFull"] == item["parent_path"]]
                        if len(pairList) > 0:
                            for pair in pairList:
                                parentWk(pair,processdict,parentPairs,line,matches,tmpFunctionList,line_number,callFunc,retDist)                            

        #Function-Function
        colNum=5
        filtered_data_header = [item for item in importList_header if str.upper(callFunc["fileNameFull"]) == str.upper(item["ParentPath"]+"\\"+item['fileName'])]        

        for data_header in filtered_data_header:
            
            
            if data_header["line"].startswith("//") or data_header["line"].startswith("/*"):
                continue
            calleeKey = ".".join(data_header["line"].replace("import","").replace(";","").strip().split(".")[:colNum])
            tmpFunctionList = [item for item in processdict[calleeKey] if (str.upper(item['fileName']) == str.upper(data_header['funcition']) or str.upper(item['fileName']) == str.upper(data_header['funcition'] + "Impl"))]

            function_pattern = "|".join(map(re.escape, [item['function'] for item in tmpFunctionList]))


            filtered_data_detail = [item for item in importList_detail 
                                        if str.upper(item['fileName'])==str.upper(data_header['fileName']) and str.upper(item["ParentPath"])==str.upper(data_header["ParentPath"]) and str.upper(item['funcition'])==str.upper(data_header['funcition'])
                                        and not(item["line"].strip().startswith("//")) and not(item["line"].strip().startswith("/*"))]     

            for data_detail in filtered_data_detail:
                matches = extract_nested_functions(data_detail["line"],function_pattern)            

                # if tmpFileName == "EDWCustomerInput":
                #     print()

                if len(matches) > 0 and len(tmpFunctionList) > 0:
                    
                    for match in matches:
        
                        callee_function_name = ""
                        callee_instance_name =  "" 
                        callee_class_name = ""
                        caller_function_name = ""
        
                        callee_instance_name, callee_function_name = (match[0],match[1])
                        callee_class_name = [item['fileName'] for item in tmpFunctionList if item['function'] == callee_function_name][0]        
        
                        #Check Caller Func
                        for callerFunction in  [item for item in targetProcess if str.upper(item['fileName']) == str.upper(callFunc['fileName']) and str.upper(item["fileNameFull"]) == str.upper(callFunc["fileNameFull"])] :                   
                            if int(callerFunction["startNum"]) <= int(data_detail["colNum"]) <= int(callerFunction["endNum"]):
                                caller_function_name = callerFunction['function']
                                break
                        
                        if caller_function_name == "":
                            continue

                        parentKey =  str.upper(callFunc["fileNameFull"] + "_" +caller_function_name)
                        childKey = str.upper([item["fileNameFull"] for item in tmpFunctionList if str.upper(item['function']) == str.upper(callee_function_name)][0] + "_" + callee_function_name)
                        
                        if (parentKey,childKey) in retDist :
                            continue
                        
                        retDist[(parentKey,childKey)] = [caller_function_name+"_"+ callFunc['fileName'], callee_function_name+"_"+ callee_class_name,False] #0:Function / 1:SQL
                        print(parentKey+","+childKey+","+caller_function_name+"_"+ callFunc['fileName']+","+callee_function_name+"_"+ callee_class_name)
                else:
                    continue
                    # parentKey = callFunc['function'] + "_" + callFunc["fileNameFull"]
                    # childKey = "None"
                    # #retDist[(parentKey,childKey)] = None
                    # print(parentKey+",'',"+ callFunc['function']+"_"+ callFunc['fileName']+",''")
    print(f"End call : {file_path}") 
    return retDist

def writeItem(processDict,resultList,importList_SQL):

    #Excel出力
    fileNo = 0
    procEndList = []

    while(True):

        fileNo += 1
        outputFile =  "./output" + getTimeString() + "_" + str(fileNo) + ".xlsx"

        # 新しいワークブックを作成
        wb = Workbook()

        # アクティブなシートを取得
        ws = wb.active

        # データを追加
        ws.title = "イベントシート"
        
        calRow=1
        calCol=1
        
        ws.cell(row=calRow, column=calCol, value=f"対象ファイル")
        target_column_letter = get_column_letter(calCol)  # 数字をアルファベットに変換
        ws.column_dimensions[target_column_letter].width = 25
        calCol+=1

        for i in range(15):
            target_column_letter = get_column_letter(calCol)  
            ws.column_dimensions[target_column_letter].width = 50
            ws.cell(row=calRow, column=calCol, value=f"区分{calCol-1}")
            calCol+=1

        #dicon/SQL
        ws.column_dimensions[get_column_letter(SQLID_COL)].width = 50
        ws.column_dimensions[get_column_letter(SQL_COL)].width = 50
        ws.cell(row=calRow, column=SQLID_COL, value=f"SQLID")
        ws.cell(row=calRow, column=SQL_COL, value=f"SQL")

        tmpClass = ""
        calCol+=1
            
        for processKey,processValue in processDict.items():

            sqlTargetList = [] 
            
            calRow+=1
            calCol=1

            if tmpClass != processValue['fileName']:
                ws.cell(row=calRow, column=calCol, value=f"{processValue['fileName']}")
                tmpClass = processValue['fileName']            
            print(f"col:{calCol}/row:{calRow} {processValue['function']}")
            calCol+=1

            # if tmpClass == "TEMAKSummaryCreate.java":
            #     print()

            ws.cell(row=calRow, column=calCol, value=f"{processValue['function']}")
            calRow = writeItemRecusively(ws,calRow,calCol,processKey,resultList,importList_SQL,processValue['function'],sqlTargetList)

            if len(sqlTargetList) > 0:
                ws.cell(row=calRow, column=1, value="SQL_NG : " + ",".join(sqlTargetList))

            sqlTargetList.clear()

            if calRow > 500000:
                break

            procEndList.append(processKey)

        # 全体のフォントを Meiryo に設定
        meiryo_font = Font(name="Meiryo")
        for row in ws.iter_rows():
            for cell in row:
                cell.font = meiryo_font

        # ファイルに保存
        wb.save(f"{outputFile}")

        print(f"{outputFile}のファイル作成が完了しました！")
        
        for procEnd in procEndList:            
            del processDict[procEnd]
        
        if len(processDict) == 0:
            break

        procEndList.clear()

def writeItemRecusively(ws,calRow,calCol,processKey,resultList,importList_SQL,exValue,sqlTargetList):

        filteredDict = {tkey:tvalue for tkey,tvalue in resultList.items() if tkey[0] == processKey}
        tmpCalRow=calRow
        cnt = 0

        if len(filteredDict)==0 :
            return(tmpCalRow)
        else:
            
            tmpCalCol=calCol + 1
            recCountMax = len(filteredDict)
            for key,value in filteredDict.items():
                
                # if value[1] == "getDeleteSql_TEMAKSummaryCreateDao":
                #     print()

                exKeyList = exValue.split("$")
                if len(exKeyList) > 0:
                    if value[1] in exKeyList:
                        continue
                    else:
                        exValue += "$"+value[1]                

                cnt+=1
                ws.cell(row=tmpCalRow, column=tmpCalCol, value=f"{value[1]}")

                print(f"col:{tmpCalCol}/row:{tmpCalRow} {value[1]}")
                if value[2] == True:
                    #dicon/SQL
                    ws.cell(row=tmpCalRow, column=SQLID_COL, value=value[1])
                    ws.cell(row=tmpCalRow, column=SQL_COL, value=[item["SQL"] for item in importList_SQL if item['funcition'] == value[1]][0])

                    for sql_Keywords in SQL_TARGETS_LIST:
                        
                        if sql_Keywords in [item["SQL"] for item in importList_SQL if item['funcition'] == value[1]][0]:
                            sqlTargetList.append(sql_Keywords)

                ret = writeItemRecusively(ws,tmpCalRow,tmpCalCol,key[1],resultList,importList_SQL,exValue,sqlTargetList)   

                tmpCalRow=ret+1                
                
                
        return(tmpCalRow)


def runParalell(directory_path,importList_header,importList_detail,importList_SQL):

    #runParalell

    processdict={}
    with ProcessPoolExecutor() as executor:
        futures = []
        for entry in os.scandir(directory_path):
            if entry.is_dir(): 
                file_path = entry.path+"\\output.csv"
                futures.append(executor.submit(load_csv_to_objects,file_path))

        for future in concurrent.futures.as_completed(futures):
            try:
                key_raw,value = future.result()
                key = key_raw.replace(directory_path+"\\","").split("\\")[0].replace("_",".")
                processdict[key] = value
            except Exception as e:
                print(f"Error processing folder: {e}") 


    resultList = {}

    # targetURL = "c:\\app\\extractModule\\output\\list\\jp_co_komatsu_emdw_web"
    # for entry in os.scandir(directory_path):

    #     if entry.is_dir(): 
    #         file_path = entry.path+"\\output.csv"
    #         target = entry.path.split("\\")[-1].replace("_",".")

    #         # if targetURL != entry.path:
    #         #     continue

    #         resultList.update(call(file_path,target,processdict,importList_header,importList_detail))

    with ProcessPoolExecutor() as executor:
        futures = []
        for entry in os.scandir(directory_path):
            if entry.is_dir(): 
                file_path = entry.path+"\\output.csv"                
                target = entry.path.split("\\")[-1].replace("_",".")
                parent_path = entry.path+"\\parents.csv"
                futures.append(executor.submit(call,file_path,target,processdict,importList_header,importList_detail,importList_SQL,parent_path))

    for future in concurrent.futures.as_completed(futures):
        try:
            resultList.update(future.result())                            
        except Exception as e:
            print(f"Error processing folder: {e}") 

    # for resultKey,resultValue in resultList.items():
    #     if any(resultKey[0] == key[1] for key in resultList.keys()):
    #         resultList[resultKey][2] = True
    
    print(f"Start create document : {file_path}") 
    output_file = crrDir +  "\\output_TMP.csv"
    processDict = {}
    with open(output_file, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(["callerFull","calleeFull","caller","callee","isSQL"])

        for resultKey,resultValue in resultList.items():
            if not(any(resultKey[0] == key[1]  for key in resultList.keys())) and not(resultValue[2]):
                if not(resultKey[0] in processDict):
                    processDict[resultKey[0]] = {"function":resultValue[0],"fileName":resultKey[0].split("\\")[-1].split("_")[0]}
                    # processDict[resultKey[0]] = {"function":resultValue[0],"fileName":resultValue[0].split("_")[-1]+".java"}
            writer.writerow([resultKey[0],resultKey[1],resultValue[0],resultValue[1],resultValue[2]])
            #resultList[resultKey][2] = True                               
        processDict = OrderedDict(sorted(processDict.items()))
        

    writeItem(processDict,resultList,importList_SQL)

if __name__ == "__main__":

    # tmpOutput_dir = crrDir + "\\output\\list\\"
    # if os.path.isdir(tmpOutput_dir):
    #     shutil.rmtree(tmpOutput_dir)
    # if not(os.path.isdir(tmpOutput_dir)):    
    #     os.mkdir(tmpOutput_dir)
    #commonRec
    importList = load_csv_to_objects(input("セットする対象ファイルパス:"))
    importList_header = [importItem for importItem in importList[1] if importItem["header/detail"] == "h" ]
    importList_detail = [importItem for importItem in importList[1] if importItem["header/detail"] == "d" ]    
    importList_TMPSQL = load_csv_to_objects(input("セットする対象ファイルパス(SQL):"))
    importList_SQL = [importItem for importItem in importList_TMPSQL[1]]
    importList[1].clear()
    importList_TMPSQL[1].clear()

    #run('N:\\New_EQP-Care(Web)\\emd-web-struts2.5\\src\\jp\\co\\komatsu\\emdw\\business\\service\\impl',"N:\\New_EQP-Care(Web)\\emd-web-struts2.5\\src\\")
    runParalell(input("プロセスフォルダ:"),importList_header,importList_detail,importList_SQL)
    #runParalell("N:\\New_EQP-Care(Web)\\emd-web-struts2.5\\src\\jp\\co\\komatsu\\emdw\\")
